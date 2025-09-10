import json
import os
from typing import List, Optional

import xlrd
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from backend.models.employee import Employee as EmployeeModel

from ..department.models import DepartmentCreate
from ..department.service import create_department, get_department_by_name
from .models import EmployeeCreate, EmployeeInDB, EmployeeUpdate


async def process_file(file):
    """处理上传的员工文件"""
    # 保存上传的文件到临时位置
    file_location = f"temp_{file.filename}"
    with open(file_location, "wb+") as file_object:
        file_object.write(await file.read())

    # 根据文件扩展名选择处理方式
    employees_data = []
    if file.filename.endswith(".xls"):
        employees_data = await process_xls_file(file_location)
    elif file.filename.endswith(".xlsx"):
        employees_data = await process_xlsx_file(file_location)

    # 处理员工数据并保存到数据库
    from config.database import SessionLocal

    db = SessionLocal()
    try:
        created_employees = []
        for emp_data in employees_data:
            # 检查部门是否存在，如果不存在则创建
            department_name = emp_data["部门"]
            department = get_department_by_name(department_name, db)
            if not department:
                # 创建新部门（使用默认值）
                dept_create = DepartmentCreate(
                    name=department_name,
                    manager="待定",
                    description=f"自动创建的部门: {department_name}",
                )
                department = create_department(dept_create, db)

            # 创建员工记录
            employee_create = EmployeeCreate(
                name=emp_data["姓名"],
                department=emp_data["部门"],
                position=emp_data["职位"],
                email=f"{emp_data['姓名'].replace(' ', '')}@example.com",  # 简单处理邮箱
                phone="",
            )

            # 保存到数据库
            db_employee = EmployeeModel(**employee_create.dict())
            db.add(db_employee)
            created_employees.append(db_employee)

        db.commit()
        # 刷新以获取ID
        for emp in created_employees:
            db.refresh(emp)

        # 转换为返回格式
        result = [
            EmployeeInDB(
                id=emp.id,
                name=emp.name,
                department=emp.department,
                position=emp.position,
                email=emp.email,
                phone=emp.phone,
            )
            for emp in created_employees
        ]

        return result
    except Exception as e:
        db.rollback()
        raise e
    finally:
        db.close()
        # 删除临时文件
        os.remove(file_location)


async def process_xls_file(file_path: str):
    """处理.xls文件"""
    workbook = xlrd.open_workbook(file_path)
    sheet = workbook.sheet_by_index(0)  # 读取第一个工作表

    # 获取表头
    headers = [str(cell.value) for cell in sheet.row(0)]

    # 验证必需的列是否存在
    required_columns = ["姓名", "职位", "部门"]
    missing_columns = [col for col in required_columns if col not in headers]
    if missing_columns:
        raise ValueError(f"缺少必需的列: {', '.join(missing_columns)}")

    # 解析数据行
    employees_data = []
    for row_idx in range(1, sheet.nrows):
        row = sheet.row(row_idx)
        employee_dict = {}
        for col_idx, header in enumerate(headers):
            if header in ["姓名", "职位", "部门"]:
                employee_dict[header] = str(row[col_idx].value).strip()
        employees_data.append(employee_dict)

    return employees_data


async def process_xlsx_file(file_path: str):
    """处理.xlsx文件"""
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # 获取表头
    headers = [str(cell.value) for cell in sheet[1]]  # 第一行是表头

    # 验证必需的列是否存在
    required_columns = ["姓名", "职位", "部门"]
    missing_columns = [col for col in required_columns if col not in headers]
    if missing_columns:
        raise ValueError(f"缺少必需的列: {', '.join(missing_columns)}")

    # 解析数据行
    employees_data = []
    for row in sheet.iter_rows(min_row=2):  # 从第二行开始是数据
        employee_dict = {}
        for col_idx, header in enumerate(headers):
            if header in ["姓名", "职位", "部门"]:
                employee_dict[header] = str(row[col_idx].value).strip()
        employees_data.append(employee_dict)

    return employees_data


def create_employee(employee_create: EmployeeCreate, db: Session) -> EmployeeInDB:
    """创建新员工"""
    # 创建员工记录
    db_employee = EmployeeModel(**employee_create.dict())

    # 保存到数据库
    db.add(db_employee)
    db.commit()
    db.refresh(db_employee)

    return EmployeeInDB(
        id=db_employee.id,
        name=db_employee.name,
        department=db_employee.department,
        position=db_employee.position,
        email=db_employee.email,
        phone=db_employee.phone,
    )


def get_employee(employee_id: int, db: Session) -> Optional[EmployeeInDB]:
    """获取指定ID的员工"""
    # 这里应该从数据库查询员工
    # 暂时返回模拟数据
    employee_data = {
        "id": employee_id,
        "name": "张三",
        "department": "技术部",
        "position": "软件工程师",
        "email": "zhangsan@example.com",
    }
    return EmployeeInDB(**employee_data)


def get_employees(
    db: Session,
    skip: int = 0,
    limit: int = 100,
) -> List[EmployeeInDB]:
    """获取员工列表"""
    # 这里应该从数据库查询员工列表
    # 暂时返回模拟数据
    employee_data = {
        "id": 1,
        "name": "张三",
        "department": "技术部",
        "position": "软件工程师",
        "email": "zhangsan@example.com",
    }
    return [EmployeeInDB(**employee_data)]


def update_employee(
    employee_id: int, employee_update: EmployeeUpdate, db: Session
) -> Optional[EmployeeInDB]:
    """更新员工"""
    # 这里应该更新数据库中的员工信息
    # 暂时返回模拟数据
    update_data = employee_update.dict(exclude_unset=True)
    employee_data = {
        "id": employee_id,
        "name": update_data.get("name", "张三"),
        "department": update_data.get("department", "技术部"),
        "position": update_data.get("position", "软件工程师"),
        "email": update_data.get("email", "zhangsan@example.com"),
    }
    return EmployeeInDB(**employee_data)


def delete_employee(employee_id: int, db: Session) -> bool:
    """删除员工"""
    # 这里应该从数据库删除员工
    # 暂时返回模拟结果
    return True
